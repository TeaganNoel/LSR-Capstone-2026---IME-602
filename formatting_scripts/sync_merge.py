#!/usr/bin/env python3

#Author: Mason Allen
#Description: Made for use in GUI. Merges dataq and mychron files into a csv
#Created Date: 4/6/2026
#Last Updated Date: 4/6/2026

"""
sync_merge.py

Purpose
-------
Merge AiM MyChron (10 Hz) and DataQ (higher rate) logs into a single Excel workbook
with one sheet per segment.

Segment definition and synchronization
--------------------------------------
- Segment count and segment durations come from the AiM file (segments detected via
  time resets to 0.0 in the AiM time column; already rounded to 0.1 s in aim parser).
- DataQ contains a channel named "Sync Volt" (combined header string) that is normally
  ~10 V, with downward spikes indicating segment initialization.
- A valid segment initialization spike is defined as:
    * Falling edge crosses below 9.5 V (threshold crossing)
    * Below-threshold maintained for at least 0.025 s and less than 2.0 s
    * Ignore any additional threshold edges for 2.0 s after a detected falling edge
      (debounce / refractory window)

The number of detected DataQ spikes must match the AiM segment count. If not, the
merge aborts (reports details and exits).

Resampling / downsampling
-------------------------
- Output uses the AiM time column for each segment (10 Hz, 0.1 s resolution).
- For each segment i:
    * Let t0_dq be the DataQ spike falling-edge time for that segment.
    * Let t_aim be the AiM time vector for that segment (starts at 0).
    * Target DataQ sampling times are: t_target = t0_dq + t_aim
- Numeric DataQ channels are low-pass filtered via a simple moving-average smoother
  before interpolation (to reduce aliasing), then interpolated onto t_target.

Special handling for DataQ non-numeric columns
----------------------------------------------
- If DataQ contains a "Date" column and/or a wall-clock "Time" column (NOT "Time ... sec"),
  these are preserved and resampled by **nearest-row selection** (no filtering, no interpolation).
- If Date values use a 2-digit year, they are normalized to 4-digit year assuming 20xx.
- Any other DataQ column that is detected as mostly non-numeric is **reported and dropped**.

Excel formatting
----------------
- Time column is written with a number format of "0.0" (always one digit after decimal).
- Numeric column formatting based on header substrings:
    * 'rpm'      -> 0.0
    * speed      -> 0.00   (matches 'speed', 'mph', 'kph')
    * temp       -> 0.0    (matches 'temp', 'deg', '°', 'egt', 'cht')
    * pressure   -> 0.000  (matches 'press', 'kpa', 'psi')  [interpreted as 0.001 kPa resolution]
    * lambda     -> 0.00
- Column widths:
    * Date column width is forced to 10 if present
    * Others: len(header)+2 clamped to [8, 60]

Dependencies
------------
- openpyxl (for writing .xlsx)
- numpy (for filtering + interpolation)
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Literal

import csv
import re

import openpyxl
from openpyxl.utils import get_column_letter

try:
    import numpy as np
except Exception as e:  # pragma: no cover
    raise SystemExit(f"ERROR: numpy is required for sync_merge.py: {e}")

from formatting_scripts import aim_parse_segments_GUI as aim
from formatting_scripts import dataq_parse_GUI as dq


# ----------------------------- Robust CSV reading ----------------------------- #

def _read_csv_table_robust(path: Path, sniff_bytes: int = 65536) -> List[List[str]]:
    """
    Read CSV into list-of-rows, but decode bytes robustly to preserve symbols (e.g., degree sign).

    Decode strategy:
      1) utf-8-sig
      2) utf-8
      3) cp1252 (common on Windows exports)
      4) latin-1 (last resort)
    """
    raw = path.read_bytes()

    last_exc: Optional[Exception] = None
    text_full = ""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text_full = raw.decode(enc, errors="strict")
            break
        except Exception as e:
            last_exc = e
            text_full = ""

    if not text_full:
        raise ValueError(f"Could not decode CSV bytes for {path}: {last_exc}")

    sample = text_full[:sniff_bytes]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    rows: List[List[str]] = []
    reader = csv.reader(text_full.splitlines(), dialect)
    for r in reader:
        rows.append([c for c in r])
    return rows


# ----------------------------- Data containers ----------------------------- #

@dataclass(frozen=True)
class TimeSeriesTable:
    """
    Generic time-series table.

    - headers: list of combined header strings aligned with columns 0..ncols-1
    - row_1:   1-based row index in original CSV for each sample
    - t:       time in seconds (float)
    - y:       2D array, shape (n_samples, ncols-1), for columns 1..ncols-1
    - raw_text: map from header index k (0..ncols-1) to an array of strings
    """
    headers: List[str]
    row_1: np.ndarray
    t: np.ndarray
    y: np.ndarray
    raw_text: Dict[int, np.ndarray]


# ----------------------------- Loading helpers ----------------------------- #

def _is_date_header(label: str) -> bool:
    lab = (label or "").strip().lower()
    return lab == "date" or lab.startswith("date ")


def _is_wallclock_time_header(label: str) -> bool:
    """
    True for a wall-clock Time column (text), false for "Time ... sec" (numeric).
    """
    lab = (label or "").strip().lower()
    if lab == "time" or lab.startswith("time "):
        return "sec" not in lab
    return False


_DATE_2Y_RE = re.compile(r"^(\s*\d{1,2}[/-]\d{1,2}[/-])(\d{2})(\s*)$")

def _normalize_date_20xx(s: str) -> str:
    """
    If s looks like M/D/YY or MM/DD/YY (or with '-' separators), convert YY -> 20YY.
    Otherwise return s unchanged.
    """
    if s is None:
        return ""
    t = str(s).strip()
    if t == "":
        return t
    m = _DATE_2Y_RE.match(t)
    if not m:
        return t
    prefix, yy, suffix = m.group(1), m.group(2), m.group(3)
    return f"{prefix}20{yy}{suffix}".strip()


def _load_aim_table(aim_csv: Path) -> TimeSeriesTable:
    """
    Load the AiM time-series block as numeric arrays.

    Respects AiM convention: data starts at (sec_row + 3).
    Time values are rounded to 0.1 s.
    """
    rows = _read_csv_table_robust(aim_csv)
    sec_row0 = aim.find_row_by_colA_value(rows, "sec")
    if sec_row0 is None:
        raise ValueError("AiM: could not find 'sec' row in column A.")

    ncols = aim.detect_ncols_from_sec_row(rows, sec_row0)
    combined = aim.parse_combined_headers(rows, sec_row0, ncols=ncols)
    headers = [h for (_col, h) in combined]

    data_start0 = sec_row0 + 3

    row_1: List[int] = []
    t_list: List[float] = []
    y_list: List[List[float]] = []

    started = False
    for r0 in range(data_start0, len(rows)):
        t_cell = aim.get_cell(rows, r0, 0)
        t = aim.parse_float_maybe(t_cell)
        if t is None:
            if started:
                break
            else:
                continue
        started = True

        t_r = aim.round_to_0p1(float(t))

        vals: List[float] = []
        for c in range(1, ncols):
            v = aim.parse_float_maybe(aim.get_cell(rows, r0, c))
            vals.append(float(v) if v is not None else float("nan"))

        row_1.append(r0 + 1)
        t_list.append(t_r)
        y_list.append(vals)

    if not row_1:
        raise ValueError("AiM: no numeric time-series rows found.")

    return TimeSeriesTable(
        headers=headers,
        row_1=np.asarray(row_1, dtype=int),
        t=np.asarray(t_list, dtype=float),
        y=np.asarray(y_list, dtype=float),
        raw_text={},
    )


def _load_dataq_table(dataq_csv: Path) -> TimeSeriesTable:
    """
    Load the DataQ time-series block as numeric arrays plus optional raw text columns.
    """
    rows = _read_csv_table_robust(dataq_csv)
    sec_row0 = dq.find_row_by_colA_value(rows, "sec")
    if sec_row0 is None:
        raise ValueError("DataQ: could not find 'sec' row in column A.")

    ncols = dq.detect_ncols_from_sec_row(rows, sec_row0)
    combined = dq.parse_combined_headers(rows, sec_row0, ncols=ncols)
    headers = [h for (_col, h) in combined]

    text_header_idxs = {
        k for k, h in enumerate(headers)
        if (k != 0) and (_is_date_header(h) or _is_wallclock_time_header(h))
    }

    r0 = sec_row0 + 1
    while r0 < len(rows):
        t = dq.parse_float_maybe(rows[r0][0] if rows[r0] else "")
        if t is not None:
            break
        r0 += 1

    if r0 >= len(rows):
        raise ValueError("DataQ: no numeric time-series rows found after 'sec' row.")

    row_1: List[int] = []
    t_list: List[float] = []
    y_list: List[List[float]] = []

    raw_text: Dict[int, List[str]] = {k: [] for k in text_header_idxs}

    for k_row in range(r0, len(rows)):
        t = dq.parse_float_maybe(rows[k_row][0] if rows[k_row] else "")
        if t is None:
            break

        vals: List[float] = []
        for c in range(1, ncols):
            cell = rows[k_row][c] if (rows[k_row] and c < len(rows[k_row])) else ""

            if c in text_header_idxs:
                s = str(cell)
                if _is_date_header(headers[c]):
                    s = _normalize_date_20xx(s)
                raw_text[c].append(s)
                vals.append(float("nan"))
            else:
                v = dq.parse_float_maybe(cell)
                vals.append(float(v) if v is not None else float("nan"))

        row_1.append(k_row + 1)
        t_list.append(float(t))
        y_list.append(vals)

        for c in text_header_idxs:
            if len(raw_text[c]) < len(row_1):
                raw_text[c].append("")

    raw_text_np: Dict[int, np.ndarray] = {k: np.asarray(v, dtype=object) for k, v in raw_text.items()}

    return TimeSeriesTable(
        headers=headers,
        row_1=np.asarray(row_1, dtype=int),
        t=np.asarray(t_list, dtype=float),
        y=np.asarray(y_list, dtype=float),
        raw_text=raw_text_np,
    )


# ----------------------------- Sync spike detection ----------------------------- #

def _find_column_index(headers: List[str], needle: str) -> Optional[int]:
    n = needle.strip().lower()
    for i, h in enumerate(headers):
        if n in (h or "").strip().lower():
            return i
    return None


def detect_sync_spikes(
    t: np.ndarray,
    sync_v: np.ndarray,
    threshold_v: float = 9.5,
    min_dur_s: float = 0.025,
    max_dur_s: float = 2.0,
    refractory_s: float = 2.0,
) -> List[Tuple[float, float, float]]:
    """
    Falling-edge detect into below-threshold region, require below-threshold maintained.

    See module docstring for spec.
    """
    if len(t) == 0:
        return []

    spikes: List[Tuple[float, float, float]] = []
    n = len(t)

    i = 1
    while i < n:
        if not (sync_v[i] < threshold_v and sync_v[i - 1] >= threshold_v):
            i += 1
            continue

        t_start = float(t[i])

        j = i
        while j < n and (sync_v[j] < threshold_v):
            j += 1

        if j < n:
            dur = float(t[j]) - t_start
            t_end = float(t[j - 1])
        else:
            dur = float(t[-1]) - t_start
            t_end = float(t[-1])

        if (dur >= min_dur_s) and (dur < max_dur_s):
            spikes.append((t_start, t_end, dur))

        t_rearm = t_start + refractory_s
        k = int(np.searchsorted(t, t_rearm, side="left"))
        i = max(j, k)
        if i < 1:
            i = 1

    return spikes


# ----------------------------- Filtering + resampling ----------------------------- #

def _moving_average(x: np.ndarray, win: int) -> np.ndarray:
    if win <= 1:
        return x
    kernel = np.ones(win, dtype=float) / float(win)
    return np.convolve(x, kernel, mode="same")


def resample_filtered(
    t: np.ndarray,
    x: np.ndarray,
    t_target: np.ndarray,
    pad_s: float = 1.0,
    ma_window_s: float = 0.1,
) -> np.ndarray:
    t0 = float(t_target[0] - pad_s)
    t1 = float(t_target[-1] + pad_s)

    i0 = int(np.searchsorted(t, t0, side="left"))
    i1 = int(np.searchsorted(t, t1, side="right"))
    i0 = max(0, min(i0, len(t) - 1))
    i1 = max(i0 + 1, min(i1, len(t)))

    tt = t[i0:i1]
    xx = x[i0:i1]

    dts = np.diff(tt)
    dt = float(np.median(dts[dts > 0])) if np.any(dts > 0) else 0.0

    win = int(round(ma_window_s / dt)) if dt > 0 else 1
    if win < 1:
        win = 1
    if win % 2 == 0:
        win += 1

    xf = _moving_average(xx, win)
    xf = _moving_average(xf, win)

    return np.interp(t_target, tt, xf)


def nearest_sample_indices(t: np.ndarray, t_target: np.ndarray) -> np.ndarray:
    pos = np.searchsorted(t, t_target, side="left")
    pos = np.clip(pos, 0, len(t) - 1)
    pos0 = np.clip(pos - 1, 0, len(t) - 1)

    d0 = np.abs(t_target - t[pos0])
    d1 = np.abs(t[pos] - t_target)
    return np.where(d0 <= d1, pos0, pos).astype(int)


# ----------------------------- Excel formatting helpers ----------------------------- #

# Maintainable table: header substring rules -> Excel number formats
# Order matters (first match wins).
FORMAT_RULES: List[Tuple[str, List[str], str]] = [
    ("rpm",      ["rpm"],                     "0.0"),
    ("lambda",   ["lambda"],                  "0.00"),
    ("speed",    ["mph", "kph", "speed"],     "0.00"),
    ("pressure", ["kpa", "psi", "press"],     "0.000"),
    ("temp",     ["temp", "deg", "°", "egt", "cht"], "0.0"),
]

def _format_for_header(header: str) -> Optional[str]:
    """Return an Excel number format for the given header, or None for General."""
    h = (header or "").strip().lower()
    for _name, keys, fmt in FORMAT_RULES:
        if any(k in h for k in keys):
            return fmt
    return None


def _numeric_fraction(x: np.ndarray) -> float:
    if x.size == 0:
        return 0.0
    return float(np.isfinite(x).sum()) / float(x.size)


def _set_column_widths(ws: openpyxl.worksheet.worksheet.Worksheet, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        h_str = str(h)
        h_low = h_str.strip().lower()
        if h_low == "date" or h_low.startswith("date "):
            width = 10
        else:
            width = len(h_str) + 2
            width = max(8, min(width, 60))
        ws.column_dimensions[get_column_letter(i)].width = width


def _apply_number_formats(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: List[str],
    n_rows: int,
    text_col_idxs_1based: set[int],
) -> None:
    """
    Apply number formats down each numeric column.
    - Column A (AiM time) always "0.0"
    - Other columns: based on FORMAT_RULES, unless the column is text.
    """
    if n_rows <= 0:
        return

    # Column A time format
    for r in range(2, 2 + n_rows):
        ws.cell(row=r, column=1).number_format = "0.0"

    # Other columns
    for c1 in range(2, len(headers) + 1):
        if c1 in text_col_idxs_1based:
            continue
        fmt = _format_for_header(headers[c1 - 1])
        if fmt is None:
            continue
        for r in range(2, 2 + n_rows):
            ws.cell(row=r, column=c1).number_format = fmt


# ----------------------------- Merge and write ----------------------------- #

def merge_to_xlsx(
    aim_csv: Path,
    dataq_csv: Path,
    out_xlsx: Path,
    min_numeric_fraction: float = 0.90,
) -> None:
    """
    Merge AiM + DataQ into a multi-sheet workbook (one sheet per segment).
    """
    aim_result = aim.analyze_aim_csv(aim_csv)
    n_seg = len(aim_result.segments)
    if n_seg == 0:
        raise ValueError("AiM: no segments detected; cannot proceed.")

    aim_tab = _load_aim_table(aim_csv)
    dq_tab = _load_dataq_table(dataq_csv)

    sync_col_idx = _find_column_index(dq_tab.headers, "sync volt")
    if sync_col_idx is None:
        raise ValueError("DataQ: could not find a column whose header contains 'Sync Volt'.")
    if sync_col_idx == 0:
        raise ValueError("DataQ: 'Sync Volt' appears to be in column A (time). Unexpected.")

    sync_v = dq_tab.y[:, sync_col_idx - 1]
    spikes = detect_sync_spikes(dq_tab.t, sync_v)
    spike_starts = [s[0] for s in spikes]

    if len(spikes) != n_seg:
        msg = [
            "ERROR: segment count mismatch between AiM and DataQ sync spikes.",
            f"  AiM segments: {n_seg}",
            f"  DataQ spikes: {len(spikes)}",
            f"  DataQ spike starts (s): {spike_starts}",
        ]
        raise SystemExit("\n".join(msg))

    keep_dq_header_idxs: List[int] = []
    keep_dq_labels: List[str] = []
    keep_kind: List[Literal["numeric", "text"]] = []

    dropped_non_numeric: List[str] = []

    for k, label in enumerate(dq_tab.headers):
        lab = (label or "").strip().lower()

        if k == 0:
            continue  # DataQ time (sec)
        if k == sync_col_idx:
            continue  # Sync Volt (sync only)
        if lab == "nc volt" or "nc volt" in lab:
            continue  # Not connected

        if _is_date_header(label) or _is_wallclock_time_header(label):
            if k not in dq_tab.raw_text:
                dropped_non_numeric.append(dq_tab.headers[k])
                continue
            keep_dq_header_idxs.append(k)
            keep_dq_labels.append(dq_tab.headers[k])
            keep_kind.append("text")
            continue

        col_data = dq_tab.y[:, k - 1]
        if _numeric_fraction(col_data) < min_numeric_fraction:
            dropped_non_numeric.append(dq_tab.headers[k])
            continue

        keep_dq_header_idxs.append(k)
        keep_dq_labels.append(dq_tab.headers[k])
        keep_kind.append("numeric")

    if dropped_non_numeric:
        print("Dropping non-numeric DataQ columns (excluding Date/Time):")
        for name in dropped_non_numeric:
            print(f"  - {name}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _select_aim_rows(start_row_1: int, end_row_1: int) -> np.ndarray:
        return np.where((aim_tab.row_1 >= start_row_1) & (aim_tab.row_1 <= end_row_1))[0]

    aim_labels = aim_tab.headers[:]  # includes time header in col A

    # Identify which output columns are text (1-based index in output sheet)
    # DataQ text columns appear after all AiM columns.
    text_cols_out_1based: set[int] = set()

    for j, kind in enumerate(keep_kind):
        if kind == "text":
            out_col_1based = len(aim_labels) + (j + 1)
            text_cols_out_1based.add(out_col_1based)

    for i_seg, seg in enumerate(aim_result.segments, start=1):
        idx = _select_aim_rows(seg.start_row_1, seg.end_row_1)
        if idx.size == 0:
            raise ValueError(
                f"AiM: empty segment selection for segment {i_seg} rows {seg.start_row_1}..{seg.end_row_1}"
            )

        t_aim = aim_tab.t[idx]         # 0.1s rounded
        y_aim = aim_tab.y[idx, :]      # AiM channels (excluding time)

        t0_dq = spikes[i_seg - 1][0]
        t_target = t0_dq + t_aim

        near_idx = nearest_sample_indices(dq_tab.t, t_target)

        dq_cols_out: List[np.ndarray] = []
        for k_hdr, kind in zip(keep_dq_header_idxs, keep_kind):
            if kind == "numeric":
                col_data = dq_tab.y[:, k_hdr - 1]
                y_rs = resample_filtered(dq_tab.t, col_data, t_target)
                dq_cols_out.append(y_rs.astype(float))
            else:
                col_text = dq_tab.raw_text[k_hdr]
                dq_cols_out.append(col_text[near_idx].astype(object))

        ws = wb.create_sheet(title=f"Export_CSV_Seg{i_seg:02d}")

        out_headers = aim_labels + keep_dq_labels
        ws.append(out_headers)

        for r in range(len(t_aim)):
            row_out: List[object] = []
            row_out.append(float(t_aim[r]))
            row_out.extend([float(v) for v in y_aim[r]])
            for col, kind in zip(dq_cols_out, keep_kind):
                if kind == "numeric":
                    row_out.append(float(col[r]))
                else:
                    row_out.append(str(col[r]))
            ws.append(row_out)

        # Formatting
        _apply_number_formats(ws, out_headers, n_rows=len(t_aim), text_col_idxs_1based=text_cols_out_1based)
        _set_column_widths(ws, out_headers)

    wb.save(out_xlsx)


# ----------------------------- Minimal CLI ----------------------------- #

def main() -> int:
    import argparse

    p = argparse.ArgumentParser(description="Merge AiM + DataQ into multi-sheet XLSX (one sheet per segment).")
    p.add_argument("aim_csv", type=Path, help="AiM MyChron CSV export")
    p.add_argument("dataq_csv", type=Path, help="DataQ CSV export")
    p.add_argument("out_xlsx", type=Path, help="Output XLSX path")
    args = p.parse_args()

    merge_to_xlsx(args.aim_csv, args.dataq_csv, args.out_xlsx)
    print(f"Wrote: {args.out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())